import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime
from typing import Dict, Optional

STORAGE_FILE = 'user_settings.xlsx'

class UserStorage:
    def __init__(self):
        self.file_path = STORAGE_FILE
        self._init_storage()
    
    def _init_storage(self):
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "User Settings"
            ws.append(['User ID', 'Username', 'Language', 'Auto Signals', 'Timezone', 'Last Updated', 'Last Activity', 'Last Welcome'])
            wb.save(self.file_path)
        else:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            if ws['G1'].value is None:
                ws['G1'] = 'Last Activity'
                ws['H1'] = 'Last Welcome'
                wb.save(self.file_path)
    
    def get_user_settings(self, user_id: int) -> Optional[Dict]:
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[0].value == user_id:
                    return {
                        'user_id': row[0].value,
                        'username': row[1].value,
                        'language': row[2].value or 'en',
                        'auto_signals': row[3].value == 'True',
                        'timezone': row[4].value or 'UTC',
                        'last_updated': row[5].value,
                        'last_activity': row[6].value if len(row) > 6 and row[6].value else None,
                        'last_welcome': row[7].value if len(row) > 7 and row[7].value else None
                    }
            return None
        except Exception as e:
            print(f"Error reading user settings: {e}")
            return None
    
    def save_user_settings(self, user_id: int, username: str, settings: Dict):
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            
            row_to_update = None
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if row[0].value == user_id:
                    row_to_update = idx
                    break
            
            language = settings.get('language', 'en')
            auto_signals = str(settings.get('auto_signals', False))
            timezone = settings.get('timezone', 'UTC')
            last_updated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            last_activity = settings.get('last_activity', last_updated)
            last_welcome = settings.get('last_welcome')
            
            if row_to_update:
                ws[f'A{row_to_update}'] = user_id
                ws[f'B{row_to_update}'] = username
                ws[f'C{row_to_update}'] = language
                ws[f'D{row_to_update}'] = auto_signals
                ws[f'E{row_to_update}'] = timezone
                ws[f'F{row_to_update}'] = last_updated
                ws[f'G{row_to_update}'] = last_activity
                ws[f'H{row_to_update}'] = last_welcome
            else:
                ws.append([user_id, username, language, auto_signals, timezone, last_updated, last_activity, last_welcome])
            
            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"Error saving user settings: {e}")
            return False
    
    def get_all_users(self):
        """Get all users from storage"""
        users = []
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    users.append({
                        'user_id': row[0],
                        'username': row[1],
                        'language': row[2] or 'en',
                        'auto_signals': row[3] == 'True',
                        'timezone': row[4] or 'UTC'
                    })
        except Exception as e:
            print(f"Error getting all users: {e}")
        
        return users
    
    def get_all_users_with_auto_signals(self):
        users = []
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3] == 'True':
                    users.append({
                        'user_id': row[0],
                        'username': row[1],
                        'language': row[2] or 'en'
                    })
        except Exception as e:
            print(f"Error getting users: {e}")
        
        return users
    
    def update_last_activity(self, user_id: int):
        """Update user's last activity timestamp"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if row[0].value == user_id:
                    ws[f'G{idx}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    wb.save(self.file_path)
                    return True
            return False
        except Exception as e:
            print(f"Error updating last activity: {e}")
            return False
    
    def get_inactive_users(self, hours=1):
        """Get users who haven't been active for specified hours"""
        users = []
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            current_time = datetime.now()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                    
                last_activity = row[6] if len(row) > 6 and row[6] else None
                last_welcome = row[7] if len(row) > 7 and row[7] else None
                
                if last_activity:
                    if isinstance(last_activity, str):
                        last_activity = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
                    
                    time_since_activity = (current_time - last_activity).total_seconds() / 3600
                    
                    send_welcome = False
                    if not last_welcome:
                        send_welcome = time_since_activity >= hours
                    else:
                        if isinstance(last_welcome, str):
                            last_welcome = datetime.strptime(last_welcome, '%Y-%m-%d %H:%M:%S')
                        time_since_welcome = (current_time - last_welcome).total_seconds() / 3600
                        send_welcome = time_since_activity >= hours and time_since_welcome >= hours
                    
                    if send_welcome:
                        users.append({
                            'user_id': row[0],
                            'username': row[1],
                            'language': row[2] or 'en'
                        })
        except Exception as e:
            print(f"Error getting inactive users: {e}")
        
        return users
    
    def update_last_welcome(self, user_id: int):
        """Update user's last welcome timestamp"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['User Settings']
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if row[0].value == user_id:
                    ws[f'H{idx}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    wb.save(self.file_path)
                    return True
            return False
        except Exception as e:
            print(f"Error updating last welcome: {e}")
            return False
